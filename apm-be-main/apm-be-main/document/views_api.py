from rest_framework import generics, filters
from .models import ProjectDocument
from .serializers import ProjectDocumentSerializer

from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle
from django.http import HttpResponse
from .models import ProjectDocument
from django.utils import timezone


from django.http import FileResponse
import tempfile

def export_documents_to_excel(request):
    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Define column names manually
    field_names = ['Project', 'Uploader', 'Name', 'Upload Date', 'Category', 'Description', 'Created At', 'Updated At', 'Download Link']

    # Write headers to the worksheet
    ws.append(field_names[:-1])  # Exclude the 'Download Link' field

    documents = ProjectDocument.objects.all()

    # Create a named style for hyperlinks
    hyperlink_style = NamedStyle(name='hyperlink_style', font=Font(color="0000FF", underline='single'))

    # Write data to the worksheet
    for document, row_num in zip(documents, range(2, len(documents) + 2)):
        # Convert upload_date to string without timezone information
        upload_date_str = document.upload_date.strftime('%Y-%m-%d')

        # Create a list with formatted values
        row_data = [
            document.project.name if document.project else '',
            document.uploader.role if document.uploader else '',
            document.name,
            upload_date_str,
            document.category if document.category else '',
            document.description if document.description else '',
            document.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            document.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            request.build_absolute_uri(document.document_file.url),  # Download Link
        ]

        for col_num, value in enumerate(row_data):
            ws.cell(row=row_num, column=col_num + 1).value = value

        # Add a hyperlink to the 'Download Link' column
        download_link_cell = ws.cell(row=row_num, column=len(field_names)).value
        ws.cell(row=row_num, column=len(field_names)).hyperlink = download_link_cell
        ws.cell(row=row_num, column=len(field_names)).style = hyperlink_style

    # Save workbook to a temporary file
    tmp_file = tempfile.NamedTemporaryFile(delete=False)
    wb.save(tmp_file.name)
    tmp_file.close()

    # Serve the file using Django FileResponse
    response = FileResponse(open(tmp_file.name, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=documents.xlsx'

    return response


#Create dan List
class ProjectDocumentListCreate(generics.ListCreateAPIView):
    queryset = ProjectDocument.objects.all()
    serializer_class = ProjectDocumentSerializer

# Edit
class ProjectDocumentRetrieveUpdate(generics.RetrieveUpdateAPIView):
    queryset = ProjectDocument.objects.all()
    serializer_class = ProjectDocumentSerializer

# Search
class ProjectDocumentListSearch(generics.ListCreateAPIView):
    queryset = ProjectDocument.objects.all()
    serializer_class = ProjectDocumentSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['id', 'name'] 